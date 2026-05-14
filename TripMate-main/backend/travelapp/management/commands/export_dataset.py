from django.core.management.base import BaseCommand
import os
import json
import csv

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

from travelapp.models import City, Flights, Hotels, Famous


def model_to_dict(obj, fields):
    d = {}
    for f in fields:
        val = getattr(obj, f)
        # simple serialization
        if hasattr(val, 'isoformat'):
            val = val.isoformat()
        d[f] = str(val) if val is not None else None
    return d


class Command(BaseCommand):
    help = "Export dataset to CSV, JSON, MongoDB JSON, SQL and Excel"

    def add_arguments(self, parser):
        parser.add_argument('--out', default='exports', help='Output directory')

    def handle(self, *args, **options):
        outdir = options['out']
        os.makedirs(outdir, exist_ok=True)

        exports = []

        # Cities
        cities = list(City.objects.all())
        city_fields = ['id', 'city', 'bestlink', 'weekgetlinks']
        city_dicts = [model_to_dict(c, city_fields) for c in cities]
        with open(os.path.join(outdir, 'cities.json'), 'w', encoding='utf-8') as f:
            json.dump(city_dicts, f, ensure_ascii=False, indent=2)
        with open(os.path.join(outdir, 'cities.csv'), 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=city_fields)
            writer.writeheader()
            writer.writerows(city_dicts)
        exports.append('cities')

        # Flights
        flights = list(Flights.objects.all())
        flight_fields = ['id', 'source', 'destination', 'flight_num', 'company', 'eprice', 'seats', 'dept_time', 'dest_time', 'stations', 'avg_rating']
        flight_dicts = [model_to_dict(flt, flight_fields) for flt in flights]
        with open(os.path.join(outdir, 'flights.json'), 'w', encoding='utf-8') as f:
            json.dump(flight_dicts, f, ensure_ascii=False, indent=2)
        with open(os.path.join(outdir, 'flights.csv'), 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=flight_fields)
            writer.writeheader()
            writer.writerows(flight_dicts)
        # MongoDB-style JSON (array of documents)
        with open(os.path.join(outdir, 'flights_mongo.json'), 'w', encoding='utf-8') as f:
            json.dump(flight_dicts, f, ensure_ascii=False, indent=2)
        # SQL inserts
        with open(os.path.join(outdir, 'flights_inserts.sql'), 'w', encoding='utf-8') as f:
            for d in flight_dicts:
                cols = ', '.join(d.keys())
                vals = ', '.join(["NULL" if d[k] is None else "'" + str(d[k]).replace("'", "''") + "'" for k in d.keys()])
                f.write(f"INSERT INTO travelapp_flights ({cols}) VALUES ({vals});\n")
        exports.append('flights')

        # Hotels
        hotels = list(Hotels.objects.all())
        hotel_fields = ['id', 'hotel_name', 'hotel_address', 'hotel_price', 'hotel_rating', 'amenities', 'distfromap', 'rooms']
        hotel_dicts = [model_to_dict(h, hotel_fields) for h in hotels]
        with open(os.path.join(outdir, 'hotels.json'), 'w', encoding='utf-8') as f:
            json.dump(hotel_dicts, f, ensure_ascii=False, indent=2)
        with open(os.path.join(outdir, 'hotels.csv'), 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=hotel_fields)
            writer.writeheader()
            writer.writerows(hotel_dicts)
        exports.append('hotels')

        # Famous places
        places = list(Famous.objects.all())
        place_fields = ['id', 'place_name', 'desc']
        place_dicts = []
        for p in places:
            d = model_to_dict(p, place_fields)
            # include image path if present
            try:
                img = p.image.url if p.image else None
            except Exception:
                img = None
            d['image'] = img
            place_dicts.append(d)
        with open(os.path.join(outdir, 'places.json'), 'w', encoding='utf-8') as f:
            json.dump(place_dicts, f, ensure_ascii=False, indent=2)
        with open(os.path.join(outdir, 'places.csv'), 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=place_fields + ['image'])
            writer.writeheader()
            writer.writerows(place_dicts)
        exports.append('places')

        # Excel workbook if available
        excel_path = os.path.join(outdir, 'dataset.xlsx')
        if OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()
            # cities
            ws = wb.active
            ws.title = 'Cities'
            ws.append(city_fields)
            for row in city_dicts:
                ws.append([row.get(f) for f in city_fields])
            # flights
            ws = wb.create_sheet('Flights')
            ws.append(flight_fields)
            for row in flight_dicts:
                ws.append([row.get(f) for f in flight_fields])
            # hotels
            ws = wb.create_sheet('Hotels')
            ws.append(hotel_fields)
            for row in hotel_dicts:
                ws.append([row.get(f) for f in hotel_fields])
            # places
            ws = wb.create_sheet('Places')
            ws.append(place_fields + ['image'])
            for row in place_dicts:
                ws.append([row.get(f) for f in (place_fields + ['image'])])
            wb.save(excel_path)
            self.stdout.write(self.style.SUCCESS(f'Excel written: {excel_path}'))
        else:
            self.stdout.write(self.style.WARNING('openpyxl not installed; skipping Excel.'))

        self.stdout.write(self.style.SUCCESS(f'Exports written to {outdir}: {", ".join(exports)}'))
